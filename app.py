from flask import Flask, render_template, request, send_file, redirect, url_for, jsonify
from openpyxl import load_workbook
from html.parser import HTMLParser
import pandas as pd
import io
import os
import re
import json
import sys
import webbrowser
import threading
import time

# --- 0. EXE PATH HANDLING ---
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

app = Flask(__name__, template_folder=resource_path("templates"))
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# File path for permanent storage
DB_FILE = 'asset_database.json'

# --- 1. PERSISTENCE FUNCTIONS ---
def load_db():
    if os.path.exists(DB_FILE):
        try:
            with open(DB_FILE, 'r') as f:
                print(f"   [SYSTEM] Loaded database from {DB_FILE}")
                return json.load(f)
        except Exception as e:
            print(f"   [SYSTEM] Error loading DB: {e}")
    return {}

def save_db(data):
    try:
        with open(DB_FILE, 'w') as f:
            json.dump(data, f, indent=4)
        print(f"   [SYSTEM] Database saved to {DB_FILE}")
    except Exception as e:
        print(f"   [SYSTEM] Error saving DB: {e}")

# Initialize global storage
app.config['REMARKS_BY_ID'] = load_db()

# --- CONFIGURATION: AUTO-MAPPING KEYWORDS ---
AUTO_SEARCH_KEYS = {
    "Computer Name": ["Computer Name"],
    "User Name": ["user name", "username"],
    "Department": ["department"], 
    "Section/Division": ["section", "division"], 
    "PC/Laptop SN": ["PC/Laptop SN:"],
    "HD SN": ["HD sn"],
    "RAM1 SN": ["Ram 1 sn:"],
    "Wireless Ethernet Address": ["Wireless Ethernet Address"],
    "Mouse S/N": ["Mouse Brand/SN:"],
    "Lock S/N": ["Notebook Lock Brand/SN:"],
    "USB-C to RJ45 Gigabit S/N": ["USB-C to RJ45 Gigabit/SN:"],
    "Power Adaptor S/N": ["Power Adaptor SN:"]
}

MANUAL_MAPPING = {
    "Region": "C3", "Location/Division": "C4", "Computer Name": "F10",
    "PC/Laptop SN": "C6", "HD SN": "C12", "RAM1 SN": "C21",
    "Wireless Ethernet Address": "F21", "User Name": "F6",
    "Mouse S/N": "C18", "Lock S/N": "C15", "USB-C to RJ45 Gigabit S/N": "C16",
    "Power Adaptor S/N": "F26", "OSSM Ticket No": "F3", "Delivery Note No": "F4",
}

# --------------------- TEXT PARSER ---------------------
def parse_comment_text(comment):
    if not comment: return {}
    lines = [line.strip() for line in comment.splitlines() if line.strip()]
    full_text = " ".join(lines)
    def extract(keyword, text):
        pattern = rf'{keyword}[\s:\-]*([A-Z0-9\-\.]{{3,}})'
        match = re.search(pattern, text, re.IGNORECASE)
        return match.group(1).strip() if match else ""
    return {
        "Mouse S/N": extract("mouse", full_text),
        "Lock S/N": extract("lock", full_text),
        "USB-C to RJ45 Gigabit S/N": extract("rj45", full_text),
        "Power Adaptor S/N": extract(r"adapt[eo]r", full_text)
    }

# --------------------- HTML PARSER ---------------------
class TableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.tables, self.curr_table, self.curr_row, self.curr_cell = [], [], [], ""
        self.in_td = self.in_tr = self.in_table = False

    def handle_starttag(self, tag, attrs):
        if tag == "table": self.in_table, self.curr_table = True, []
        elif tag == "tr" and self.in_table: self.in_tr, self.curr_row = True, []
        elif tag == "td" and self.in_tr: self.in_td, self.curr_cell = True, ""

    def handle_endtag(self, tag):
        if tag == "td" and self.in_td:
            self.curr_row.append(self.curr_cell.strip())
            self.in_td = False
        elif tag == "tr" and self.in_tr:
            if self.curr_row: self.curr_table.append(self.curr_row)
            self.in_tr = False
        elif tag == "table" and self.in_table:
            if self.curr_table: self.tables.append(self.curr_table)
            self.in_table = False

    def handle_data(self, data):
        if self.in_td: self.curr_cell += data

def extract_hardware_from_html(html):
    # Standardize HTML: remove the hidden span "dots" that break parsing
    clean_html = re.sub(r'<span[^>]*>\.</span>', '', html)
    
    parser = TableParser()
    parser.feed(clean_html)
    
    all_keys = set(AUTO_SEARCH_KEYS.keys()) | set(MANUAL_MAPPING.keys())
    data = {k: "" for k in all_keys}
    
    # --- 1. COMPUTER NAME ---
    # Using Regex for maximum reliability
    name_match = re.search(r'<td>CSName</td>\s*<td>([^<]+)', clean_html, re.IGNORECASE)
    if name_match:
        data["Computer Name"] = name_match.group(1).replace('.', '').strip()

    # --- 2. WIRELESS ETHERNET (Intel Wi-Fi) ---
    # Find the table row that contains "Intel" and a MAC Address
    mac_pattern = r'<td>(Intel\(R\) Wi-Fi[^<]+)</td>\s*<td>([0-9A-F:]{17})'
    mac_match = re.search(mac_pattern, clean_html, re.IGNORECASE)
    if mac_match:
        data["Wireless Ethernet Address"] = mac_match.group(2).strip()
    else:
        # Fallback for the other table format seen in your screenshots
        for t in parser.tables:
            table_text = "".join(str(cell).lower() for row in t for cell in row)
            if "intel" in table_text and "wi-fi" in table_text:
                headers = [str(c).lower().strip() for c in t[0]]
                if "macaddress" in headers:
                    m_idx = headers.index("macaddress")
                    d_idx = headers.index("description")
                    for r in t[1:]:
                        if "intel" in str(r[d_idx]).lower():
                            data["Wireless Ethernet Address"] = str(r[m_idx]).replace('.', '').strip()

    # --- 3. HD SERIAL NUMBER (PHYSICALDRIVE0) ---
    # Targets the exact row from your screenshot: SerialNumber followed by Tag
    for t in parser.tables:
        table_text = "".join(str(cell).lower() for row in t for cell in row)
        if "physicaldrive0" in table_text:
            headers = [str(c).lower().strip() for c in t[0]]
            if "serialnumber" in headers:
                s_idx = headers.index("serialnumber")
                t_idx = headers.index("tag") if "tag" in headers else -1
                for r in t[1:]:
                    # Ensure we are looking at DRIVE 0
                    tag_val = str(r[t_idx]).upper() if t_idx != -1 else ""
                    if "PHYSICALDRIVE0" in tag_val or t_idx == -1:
                        # Clean "E0F7.." into "E0F7"
                        raw_sn = str(r[s_idx]).strip()
                        data["HD SN"] = raw_sn.split('.')[0].replace('.', '').strip()
                        break

    # --- 4. RAM SERIAL NUMBER ---
    # Extract from the Win32_PhysicalMemory table
    for t in parser.tables:
        if len(t) > 0 and any("manufacturer" in str(c).lower() for c in t[0]):
            headers = [str(c).lower() for c in t[0]]
            try:
                manu_i = headers.index("manufacturer")
                part_i = headers.index("partnumber")
                serial_i = headers.index("serialnumber")
                r = t[1]
                data["RAM Type / Capacity"] = f"{r[manu_i]} - {r[part_i]}"
                data["RAM1 SN"] = r[serial_i]
            except Exception:
                pass
            break

    # --- 5. PC SERIAL NUMBER (BIOS) ---
    bios_match = re.search(r'Win32_BIOS.*?SerialNumber.*?<td>([^<]+)</td>', clean_html, re.DOTALL | re.IGNORECASE)
    if bios_match:
        data["PC/Laptop SN"] = bios_match.group(1).replace('.', '').strip()
            
    return data

# --------------------- TEMPLATE MAPPER ---------------------
def get_template_mapping(sheet):
    mapping = MANUAL_MAPPING.copy()
    found_keys = set(mapping.keys())
    for row in sheet.iter_rows(min_row=1, max_row=50, max_col=20):
        for cell in row:
            if not cell.value: continue
            cell_text = str(cell.value).lower().strip().rstrip(":")
            for field, phrases in AUTO_SEARCH_KEYS.items():
                if field in found_keys: continue 
                if any(phrase.lower() == cell_text for phrase in phrases):
                    target_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                    mapping[field] = target_cell.coordinate
                    found_keys.add(field)
    return mapping

# --------------------- ROUTES ---------------------
@app.route('/')
def index():
    count = len(app.config['REMARKS_BY_ID'])
    return render_template('index.html', error=request.args.get('error'), message=request.args.get('message'), count=count)

@app.route('/upload_excel', methods=['POST'])
def upload_excel():
    if 'file' not in request.files: return redirect(url_for('index', error="No file uploaded"))
    file = request.files['file']
    try:
        wb = load_workbook(file, read_only=False, data_only=True)
        ws = wb["2021 NEW"] if "2021 NEW" in wb.sheetnames else wb.active
        remarks_map = app.config['REMARKS_BY_ID']
        headers = [cell.value for cell in ws[1]]
        
        def find_col(keywords):
            for i, h in enumerate(headers):
                if h and any(k in str(h).lower() for k in keywords): return i
            return -1

        dept_idx = find_col(["department", "dept"])
        sect_idx = find_col(["section", "division", "sect", "location"]) # Expanded
        user_idx = find_col(["user"])
        region_idx = 10 

        for row in ws.iter_rows(min_row=2, max_col=20): 
            id_val = row[1].value 
            if not id_val: continue
            id_str = str(id_val).strip().replace(".0", "").zfill(4)
            note = ""
            for i in range(min(3, len(row))):
                if row[i].comment:
                    note = row[i].comment.text.strip(); break 

            remarks_map[id_str] = {
                "note": note,
                "dept": str(row[dept_idx].value).strip() if dept_idx != -1 and row[dept_idx].value else "",
                "sect": str(row[sect_idx].value).strip() if sect_idx != -1 and row[sect_idx].value else "",
                "user": str(row[user_idx].value).strip() if user_idx != -1 and row[user_idx].value else "",
                "region": str(row[region_idx].value).strip() if len(row) > region_idx and row[region_idx].value else ""
            }
        save_db(remarks_map)
        return redirect(url_for('index', message=f" Saved! Database holds {len(remarks_map)} assets."))
    except Exception as e: return redirect(url_for('index', error=f"Excel Error: {str(e)}"))

@app.route('/upload_html', methods=['POST'])
def upload_html():
    if 'file' not in request.files: return redirect(url_for('index', error="No HTML file"))
    file = request.files['file']

    ossm_ticket = request.form.get('ossm_ticket', '')
    delivery_note = request.form.get('delivery_note', '')

    try:
        content = file.read().decode('utf-8', errors='ignore')
        data = extract_hardware_from_html(content)
        comp_name = data.get("Computer Name", "Unknown")
        match = re.search(r'(\d{4})$', comp_name) 

        data["OSSM Ticket No"] = ossm_ticket
        data["Delivery Note No"] = delivery_note
        
        if match:
            asset_info = app.config['REMARKS_BY_ID'].get(match.group(1))
            if asset_info:
                # Improved Section/Division mapping
                db_section = asset_info.get("sect", "")
                data["User Name"] = asset_info.get("user", "")
                data["Department"] = asset_info.get("dept", "")
                data["Region"] = asset_info.get("region", "")
                data["Section/Division"] = db_section
                data["Location/Division"] = db_section 
                
                if asset_info.get("note"): data.update(parse_comment_text(asset_info["note"]))

        template_path = resource_path('configuration_form template.xlsx')
        if not os.path.exists(template_path):
            return redirect(url_for('index', error="Template file not found!"))

        wb = load_workbook(template_path); sheet = wb.active 
        cell_map = get_template_mapping(sheet)
        for field, value in data.items():
            if field in cell_map and value: sheet[cell_map[field]] = value
        
        output = io.BytesIO(); wb.save(output); output.seek(0)
        filename = f"{comp_name}_ConfigForm.xlsx"
        return send_file(output, download_name=filename, as_attachment=True)
    except Exception as e: return redirect(url_for('index', error=f"HTML Error: {str(e)}"))



@app.route('/clear_db')
def clear_db():
    app.config['REMARKS_BY_ID'] = {}
    if os.path.exists(DB_FILE): os.remove(DB_FILE)
    return redirect(url_for('index', message=" Database wiped successfully."))

# --- STARTUP LOGIC ---
def open_browser():
    time.sleep(1.5) 
    webbrowser.open("http://127.0.0.1:5000")

if __name__ == '__main__':
    if not os.environ.get('WERKZEUG_RUN_MAIN'):
        browser_thread = threading.Thread(target=open_browser)
        browser_thread.daemon = True
        browser_thread.start()

    app.run(debug=True, port=5000)